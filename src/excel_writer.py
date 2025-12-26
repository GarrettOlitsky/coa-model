from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_coa_xlsx(accounts: list[dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    ws.append(["Account Number", "Account Name"])
    for a in accounts:
        ws.append([a["number"], a["name"]])

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 44

    wb.save(out_path)
